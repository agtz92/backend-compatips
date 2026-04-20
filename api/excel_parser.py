"""Parsea archivos Excel de facturación y movimientos bancarios.

Detecta encabezados de columna por palabras clave (es flexible al formato del
archivo de origen) y normaliza cada fila a un diccionario.
"""
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
import re
import unicodedata

from openpyxl import load_workbook


# Mapeo de campo lógico -> palabras clave (se buscan dentro del header)
FACTURA_HEADERS = {
    'folio': ['folio', 'no factura', 'numero factura', 'num factura', 'invoice', 'uuid', 'serie folio'],
    'fecha': ['fecha emision', 'fecha factura', 'fecha de emision', 'fecha'],
    'cliente': ['cliente', 'razon social', 'receptor', 'nombre cliente'],
    'concepto': ['concepto', 'descripcion', 'detalle'],
    'total': ['total', 'monto total', 'importe total', 'gran total'],
}

MOVIMIENTO_HEADERS = {
    'fecha': ['fecha operacion', 'fecha movimiento', 'fecha'],
    'descripcion': ['descripcion', 'concepto', 'detalle', 'movimiento'],
    'referencia': ['referencia', 'ref', 'folio'],
    'monto': ['abono', 'deposito', 'monto', 'importe', 'cargo'],
    'tipo': ['tipo', 'naturaleza'],
}


def _normalize(text):
    if text is None:
        return ''
    text = str(text).strip().lower()
    text = unicodedata.normalize('NFKD', text)
    text = ''.join(c for c in text if not unicodedata.combining(c))
    text = re.sub(r'[^a-z0-9 ]+', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def _match_header(cell_value, keywords):
    """¿El header normalizado contiene alguna keyword?"""
    norm = _normalize(cell_value)
    if not norm:
        return False
    return any(kw in norm for kw in keywords)


def _find_header_row(ws, header_specs, max_rows=15):
    """Encuentra la fila de encabezados y devuelve {campo_logico: indice_columna}.

    Recorre las primeras max_rows filas hasta encontrar la que mapee a la
    mayoría de los campos esperados.
    """
    best_row = None
    best_mapping = {}
    for row_idx in range(1, max_rows + 1):
        mapping = {}
        for col_idx, cell in enumerate(ws[row_idx], start=1):
            for field, keywords in header_specs.items():
                if field in mapping:
                    continue
                if _match_header(cell.value, keywords):
                    mapping[field] = col_idx
                    break
        if len(mapping) > len(best_mapping):
            best_mapping = mapping
            best_row = row_idx
            if len(mapping) == len(header_specs):
                break
    if not best_mapping:
        raise ValueError('No se encontraron encabezados reconocibles en la hoja.')
    return best_row, best_mapping


def _to_decimal(value):
    if value is None or value == '':
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None
    text = str(value).strip()
    text = text.replace('$', '').replace(',', '').replace(' ', '')
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _to_date(value):
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ('%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d', '%Y/%m/%d', '%d-%m-%y', '%d/%m/%y'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _to_str(value):
    if value is None:
        return ''
    return str(value).strip()


def parse_facturas(file_obj):
    """Parsea el Excel de facturación. Devuelve lista de dicts."""
    wb = load_workbook(file_obj, data_only=True, read_only=True)
    ws = wb.active
    header_row, mapping = _find_header_row(ws, FACTURA_HEADERS)

    required = {'folio', 'fecha', 'total'}
    missing = required - set(mapping.keys())
    if missing:
        raise ValueError(f'Faltan columnas requeridas en facturación: {sorted(missing)}')

    facturas = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(c is not None and str(c).strip() for c in row):
            continue
        def get(field):
            idx = mapping.get(field)
            return row[idx - 1] if idx and idx <= len(row) else None

        folio = _to_str(get('folio'))
        fecha = _to_date(get('fecha'))
        total = _to_decimal(get('total'))
        if not folio or fecha is None or total is None:
            continue

        facturas.append({
            'folio': folio,
            'fecha': fecha,
            'cliente': _to_str(get('cliente')),
            'concepto': _to_str(get('concepto')),
            'total': total,
            'fila_origen': {k: _to_str(get(k)) for k in mapping},
        })
    return facturas


def parse_movimientos(file_obj):
    """Parsea el Excel de movimientos bancarios. Devuelve lista de dicts."""
    wb = load_workbook(file_obj, data_only=True, read_only=True)
    ws = wb.active
    header_row, mapping = _find_header_row(ws, MOVIMIENTO_HEADERS)

    if 'fecha' not in mapping or 'monto' not in mapping:
        raise ValueError('El Excel de movimientos debe incluir fecha y monto.')

    movimientos = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(c is not None and str(c).strip() for c in row):
            continue
        def get(field):
            idx = mapping.get(field)
            return row[idx - 1] if idx and idx <= len(row) else None

        fecha = _to_date(get('fecha'))
        monto = _to_decimal(get('monto'))
        if fecha is None or monto is None:
            continue
        # solo considerar abonos positivos (depósitos)
        if monto <= 0:
            continue

        movimientos.append({
            'fecha': fecha,
            'descripcion': _to_str(get('descripcion')),
            'referencia': _to_str(get('referencia')),
            'monto': monto,
            'tipo': _to_str(get('tipo')),
            'fila_origen': {k: _to_str(get(k)) for k in mapping},
        })
    return movimientos


def _decode_bytes(raw):
    """Decodifica bytes probando encodings comunes en exports bancarios MX."""
    for enc in ('utf-8-sig', 'utf-8', 'cp1252', 'latin-1'):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode('latin-1', errors='replace')


def _split_row(line):
    """Divide una línea por tab si existe, si no por 2+ espacios."""
    if '\t' in line:
        return [c.strip() for c in line.split('\t')]
    return [c.strip() for c in re.split(r' {2,}', line)]


def parse_movimientos_txt(file_obj):
    """Parsea un export TSV/texto del banco con columnas:
    Día | Concepto / Referencia | cargo | Abono | Saldo

    Solo se quedan las filas con valor en `Abono` (depósitos).
    El string después del primer '/' en Concepto se usa como referencia.
    """
    raw = file_obj.read()
    if isinstance(raw, bytes):
        text = _decode_bytes(raw)
    else:
        text = raw

    lines = [ln for ln in text.splitlines() if ln.strip()]
    if not lines:
        raise ValueError('El archivo está vacío.')

    header_idx = None
    for i, ln in enumerate(lines[:10]):
        norm = _normalize(ln)
        if ('dia' in norm or 'fecha' in norm) and 'abono' in norm:
            header_idx = i
            break
    if header_idx is None:
        raise ValueError('No se encontró encabezado con Día/Abono en el archivo.')

    headers = [_normalize(c) for c in _split_row(lines[header_idx])]
    def col(*keywords):
        for i, h in enumerate(headers):
            if any(k in h for k in keywords):
                return i
        return None

    idx_fecha = col('dia', 'fecha')
    idx_concepto = col('concepto', 'referencia', 'descripcion', 'movimiento')
    idx_cargo = col('cargo', 'debito', 'retiro')
    idx_abono = col('abono', 'deposito', 'credito')
    if idx_fecha is None or idx_abono is None:
        raise ValueError('El archivo debe incluir columnas Día y Abono.')

    movimientos = []
    for ln in lines[header_idx + 1:]:
        cols = _split_row(ln)
        if len(cols) <= max(idx_fecha, idx_abono):
            continue
        fecha = _to_date(cols[idx_fecha])
        if fecha is None:
            continue
        abono_raw = cols[idx_abono] if idx_abono < len(cols) else ''
        monto = _to_decimal(abono_raw)
        if monto is None or monto <= 0:
            continue
        cargo_raw = cols[idx_cargo] if idx_cargo is not None and idx_cargo < len(cols) else ''
        if _to_decimal(cargo_raw):
            continue
        concepto_full = cols[idx_concepto] if idx_concepto is not None and idx_concepto < len(cols) else ''
        if '/' in concepto_full:
            tipo, _, referencia = concepto_full.partition('/')
            tipo = tipo.strip()
            referencia = referencia.strip()
        else:
            tipo = ''
            referencia = concepto_full.strip()
        movimientos.append({
            'fecha': fecha,
            'descripcion': concepto_full.strip(),
            'referencia': referencia,
            'monto': monto,
            'tipo': tipo,
            'fila_origen': {
                'dia': cols[idx_fecha],
                'concepto': concepto_full,
                'cargo': cargo_raw,
                'abono': abono_raw,
            },
        })
    return movimientos
